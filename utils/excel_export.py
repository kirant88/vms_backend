"""
Excel export utility for visitor data
"""

import io
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from django.http import HttpResponse
from visitors.models import Visitor


def export_visitors_to_excel(visitors, filename="visitors_export.xlsx"):
    """
    Export visitors data to Excel format
    """
    # Create workbook and worksheet
    wb = Workbook()
    ws = wb.active
    ws.title = "Visitors Data"

    # Define headers
    headers = [
        "ID",
        "Name",
        "Email",
        "Phone",
        "Company",
        "Purpose",
        "Department",
        "Visitor Type",
        "Visitor Category",
        "Host Name",
        "Visit Date",
        "Visit Time",
        "Status",
        "QR Code",
        "Created At",
        "Updated At",
    ]

    # Style definitions
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(
        start_color="366092", end_color="366092", fill_type="solid"
    )
    border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )
    center_alignment = Alignment(horizontal="center", vertical="center")

    # Add headers
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = border
        cell.alignment = center_alignment

    # Add data rows
    for row, visitor in enumerate(visitors, 2):
        data = [
            str(visitor.id),
            visitor.name,
            visitor.email,
            visitor.phone or "",
            visitor.company or "",
            visitor.purpose,
            visitor.department.name if visitor.department else "",
            visitor.visitor_type or "",
            visitor.visitor_category or "",
            visitor.host_name or "",
            str(visitor.visit_date),
            str(visitor.visit_time),
            visitor.status,
            visitor.qr_code,
            (
                visitor.created_at.strftime("%Y-%m-%d %H:%M:%S")
                if visitor.created_at
                else ""
            ),
            (
                visitor.updated_at.strftime("%Y-%m-%d %H:%M:%S")
                if visitor.updated_at
                else ""
            ),
        ]

        for col, value in enumerate(data, 1):
            cell = ws.cell(row=row, column=col, value=value)
            cell.border = border
            if col in [1, 9, 10, 13, 14]:  # ID, dates, times - center align
                cell.alignment = center_alignment

    # Auto-adjust column widths
    for col in range(1, len(headers) + 1):
        column_letter = get_column_letter(col)
        max_length = 0

        # Check header length
        header_length = len(str(headers[col - 1]))
        max_length = max(max_length, header_length)

        # Check data lengths
        for row in range(2, len(visitors) + 2):
            cell_value = ws.cell(row=row, column=col).value
            if cell_value:
                max_length = max(max_length, len(str(cell_value)))

        # Set column width (with some padding)
        adjusted_width = min(max_length + 2, 50)  # Max width of 50
        ws.column_dimensions[column_letter].width = adjusted_width

    # Create response
    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = f"attachment; filename={filename}"

    # Save workbook to response
    wb.save(response)
    return response


def export_visitors_to_excel_memory(visitors, filename="visitors_export.xlsx"):
    """
    Export visitors data to Excel format in memory (for API responses)
    """
    # Create workbook and worksheet
    wb = Workbook()
    ws = wb.active
    ws.title = "Visitors Data"

    # Define headers
    headers = [
        "ID",
        "Name",
        "Email",
        "Phone",
        "Company",
        "Purpose",
        "Department",
        "Visitor Type",
        "Visitor Category",
        "Host Name",
        "Visit Date",
        "Visit Time",
        "Status",
        "QR Code",
        "Created At",
        "Updated At",
    ]

    # Style definitions
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(
        start_color="366092", end_color="366092", fill_type="solid"
    )
    border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )
    center_alignment = Alignment(horizontal="center", vertical="center")

    # Add headers
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = border
        cell.alignment = center_alignment

    # Add data rows
    for row, visitor in enumerate(visitors, 2):
        data = [
            str(visitor.id),
            visitor.name,
            visitor.email,
            visitor.phone or "",
            visitor.company or "",
            visitor.purpose,
            visitor.department.name if visitor.department else "",
            visitor.visitor_type or "",
            visitor.visitor_category or "",
            visitor.host_name or "",
            str(visitor.visit_date),
            str(visitor.visit_time),
            visitor.status,
            visitor.qr_code,
            (
                visitor.created_at.strftime("%Y-%m-%d %H:%M:%S")
                if visitor.created_at
                else ""
            ),
            (
                visitor.updated_at.strftime("%Y-%m-%d %H:%M:%S")
                if visitor.updated_at
                else ""
            ),
        ]

        for col, value in enumerate(data, 1):
            cell = ws.cell(row=row, column=col, value=value)
            cell.border = border
            if col in [1, 9, 10, 13, 14]:  # ID, dates, times - center align
                cell.alignment = center_alignment

    # Auto-adjust column widths
    for col in range(1, len(headers) + 1):
        column_letter = get_column_letter(col)
        max_length = 0

        # Check header length
        header_length = len(str(headers[col - 1]))
        max_length = max(max_length, header_length)

        # Check data lengths
        for row in range(2, len(visitors) + 2):
            cell_value = ws.cell(row=row, column=col).value
            if cell_value:
                max_length = max(max_length, len(str(cell_value)))

        # Set column width (with some padding)
        adjusted_width = min(max_length + 2, 50)  # Max width of 50
        ws.column_dimensions[column_letter].width = adjusted_width

    # Save to memory buffer
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    return buffer.getvalue()


def get_predefined_hosts():
    """
    Get predefined hosts for the template
    """
    return [
        {"name": "Ajith Karnali (C4i4)", "email": "ajith@c4i4.org"},
        {"name": "Ankita Sharma (C4i4)", "email": "ankita.sharma@c4i4.org"},
        {"name": "Anuja Arora (C4i4)", "email": "anuja.arora@c4i4.org"},
        {"name": "Dhananjay Magdum (C4i4)", "email": "dhananjay.magdum@c4i4.org"},
        {"name": "Disha Rajpal (c4i4)", "email": "disha.rajpal@c4i4.org"},
        {"name": "Harshal Golait (C4i4)", "email": "harshal.golait@c4i4.org"},
        {"name": "Himanshu Singh(c4i4)", "email": "himanshu.singh@c4i4.org"},
        {"name": "Khushi Chaudhary (C4i4)", "email": "khushi.chaudhary@c4i4.org"},
        {"name": "Kiran Tondchore (C4i4)", "email": "kiran.tondchore@indi4.io"},
        {"name": "Natasha Acharya (c4i4)", "email": "natasha.acharya@c4i4.org"},
        {"name": "Nishkarsh Kumar (C4i4)", "email": "nishkarsh.kumar@c4i4.org"},
        {"name": "Prachi Bhorade (C4i4)", "email": "accounts@c4i4.org"},
        {"name": "Pranav Pankhawala (C4i4)", "email": "pranav.pankhawala@c4i4.org"},
        {"name": "Pranav Talekar (C4i4)", "email": "hr@c4i4.org"},
        {"name": "Rajesh Pahadi (C4i4)", "email": "rajesh.pahadi@c4i4.org"},
        {"name": "Sayali Kulkarni (C4i4)", "email": "sayali.kulkarni@c4i4.org"},
    ]


def create_bulk_visitor_template():
    """
    Create Excel template for bulk visitor registration
    """
    # Create workbook and worksheet
    wb = Workbook()
    ws = wb.active
    ws.title = "Bulk Visitor Registration Template"

    # Define headers for the template
    headers = [
        "Name*",
        "Email*",
        "Phone*",
        "Company/Organization*",
        "Visitor Type*",
        "Visitor Category*",
    ]

    # Style definitions
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(
        start_color="2E7D32", end_color="2E7D32", fill_type="solid"
    )
    border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )
    center_alignment = Alignment(horizontal="center", vertical="center")

    # Add headers
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = border
        cell.alignment = center_alignment

    # Add sample data row with instructions (using future dates)
    from datetime import datetime, timedelta

    # Get next Monday
    today = datetime.now().date()
    days_ahead = 0
    while (today + timedelta(days=days_ahead)).weekday() != 0:  # Monday is 0
        days_ahead += 1
    next_monday = today + timedelta(days=days_ahead)

    # sample_data = [
    #     "John Doe",
    #     "john.doe@example.com",
    #     "1234567890",
    #     "ABC Company",
    #     "emp",
    #     "industry",
    # ]

    # for col, value in enumerate(sample_data, 1):
    #     cell = ws.cell(row=2, column=col, value=value)
    #     cell.border = border
    #     if col in [5, 6, 7, 8, 9, 10]:  # Dropdown fields - center align
    #         cell.alignment = center_alignment

    # Add hosts sheet
    hosts_ws = wb.create_sheet("Hosts")
    hosts = get_predefined_hosts()

    # Add hosts headers
    hosts_headers = ["Host Name", "Host Email"]
    for col, header in enumerate(hosts_headers, 1):
        cell = hosts_ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = border
        cell.alignment = center_alignment

    # Add hosts data
    for row, host in enumerate(hosts, 2):
        hosts_ws.cell(row=row, column=1, value=host["name"])
        hosts_ws.cell(row=row, column=2, value=host["email"])

    # Auto-adjust column widths for hosts sheet
    hosts_ws.column_dimensions["A"].width = 25
    hosts_ws.column_dimensions["B"].width = 30

    # Add instructions sheet
    instructions_ws = wb.create_sheet("Instructions")

    instructions = [
        ["BULK VISITOR REGISTRATION TEMPLATE - INSTRUCTIONS"],
        [""],
        ["REQUIRED FIELDS (marked with *):"],
        ["• Name*: Full name of the visitor"],
        ["• Email*: Valid email address"],
        ["• Phone*: Contact phone number"],
        ["• Company/Organization: Company or organization name"],
        ["• Visitor Type*: 'emp' for Employee or 'student' for Student"],
        ["• Visitor Category*: 'industry' or 'academic'"],
    ]

    # Add department list
    from visitors.models import Department

    departments = Department.objects.all()
    for dept in departments:
        instructions.append([f"{dept.id} | {dept.name}"])

    # Add instructions to worksheet
    for row, instruction in enumerate(instructions, 1):
        cell = instructions_ws.cell(row=row, column=1, value=instruction[0])
        if row == 1:  # Title
            cell.font = Font(bold=True, size=14)
        elif instruction[0].startswith(
            ("•", "REQUIRED", "OPTIONAL", "IMPORTANT", "TEMPLATE", "DEPARTMENT")
        ):
            cell.font = Font(bold=True)

    # Auto-adjust column widths for main sheet
    for col in range(1, len(headers) + 1):
        column_letter = get_column_letter(col)
        max_length = len(headers[col - 1]) + 2
        ws.column_dimensions[column_letter].width = max_length

    # Auto-adjust column widths for instructions sheet
    instructions_ws.column_dimensions["A"].width = 80

    # Save to memory buffer
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    return buffer.getvalue()
