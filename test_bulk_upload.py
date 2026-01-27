#!/usr/bin/env python
"""
Test script for bulk upload endpoint
"""
import requests
import os
from datetime import datetime, timedelta
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment
from io import BytesIO

# Configuration
API_URL = "http://localhost:8000/api/v1/bulk/upload/"


def create_test_excel():
    """Create a test Excel file with visitor data"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Visitors"

    # Add headers
    headers = [
        "Name*",
        "Email*",
        "Phone*",
        "Company/Organization*",
        "Visitor Type*",
        "Visitor Category*",
    ]
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.value = header
        cell.fill = PatternFill(
            start_color="4472C4", end_color="4472C4", fill_type="solid"
        )
        cell.font = Font(bold=True, color="FFFFFF")
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # Add test data (3 visitors)
    test_data = [
        [
            "John Doe",
            "testkirantondchore@gmail.com",
            "1234567890",
            "ABC Company",
            "Professional",
            "Industry",
        ],
        [
            "Jon Doe",
            "kirantondchore007@gmail.com",
            "1234567891",
            "ABC pany",
            "Professional",
            "Industry",
        ],
        [
            "Jane Smith",
            "jane.smith@company.com",
            "9876543210",
            "XYZ Corp",
            "Professional",
            "Industry",
        ],
    ]

    for row_num, row_data in enumerate(test_data, 2):
        for col_num, value in enumerate(row_data, 1):
            ws.cell(row=row_num, column=col_num).value = value

    # Adjust column widths
    for col in ["A", "B", "C", "D", "E", "F"]:
        ws.column_dimensions[col].width = 20

    # Save to bytes
    excel_bytes = BytesIO()
    wb.save(excel_bytes)
    excel_bytes.seek(0)
    return excel_bytes


def test_bulk_upload():
    """Test the bulk upload endpoint"""

    # Create test Excel file
    excel_file = create_test_excel()

    # Calculate visit date (next weekday)
    visit_date = datetime.now().date() + timedelta(days=1)
    # If it's Friday, skip to Monday
    while visit_date.weekday() >= 4:  # 4 = Friday, 5 = Saturday, 6 = Sunday
        visit_date += timedelta(days=1)

    # Prepare form data
    form_data = {
        "purpose": "i_factory_visit",
        "visit_date": visit_date.strftime("%Y-%m-%d"),
        "visit_time": "11:00",
        "host_name": "Kiran Tondchore (C4i4)",
        "host_email": "kiran.tondchore@indi4.io",
    }

    # Prepare files
    files = {
        "file": (
            "bulk_visitor_template.xlsx",
            excel_file,
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        ),
    }

    print("=" * 60)
    print("BULK UPLOAD TEST")
    print("=" * 60)
    print(f"API URL: {API_URL}")
    print(f"Visit Date: {form_data['visit_date']}")
    print(f"Visit Time: {form_data['visit_time']}")
    print(f"Purpose: {form_data['purpose']}")
    print(f"Host Name: {form_data['host_name']}")
    print(f"Host Email: {form_data['host_email']}")
    print("-" * 60)

    try:
        response = requests.post(
            API_URL,
            data=form_data,
            files=files,
            timeout=30,
        )

        print(f"Status Code: {response.status_code}")
        print(f"Response Headers: {dict(response.headers)}")
        print(f"Response Body: {response.text}")

        if response.status_code == 200:
            result = response.json()
            print("\n✅ SUCCESS!")
            print(f"  - Total Processed: {result.get('total_processed')}")
            print(f"  - Successful: {result.get('successful')}")
            print(f"  - Failed: {result.get('failed')}")
            if result.get("warnings"):
                print(f"  - Warnings: {result.get('warnings')}")
            return True
        else:
            print(f"\n❌ FAILED!")
            try:
                error_data = response.json()
                if "error" in error_data:
                    print(f"  Error: {error_data['error']}")
                if "message" in error_data:
                    print(f"  Message: {error_data['message']}")
                if "errors" in error_data:
                    print(f"  Detailed Errors: {error_data['errors']}")
            except:
                print(f"  Response: {response.text}")
            return False

    except Exception as e:
        print(f"❌ Exception: {str(e)}")
        import traceback

        traceback.print_exc()
        return False


if __name__ == "__main__":
    test_bulk_upload()
