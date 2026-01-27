from django.shortcuts import get_object_or_404
from django.utils import timezone
from rest_framework import generics, status
from rest_framework.decorators import api_view, permission_classes
from rest_framework.permissions import IsAuthenticated, AllowAny
from rest_framework.response import Response
from django_filters.rest_framework import DjangoFilterBackend
from django.db.models import Q
from django.http import HttpResponse
from datetime import datetime, date, time, timedelta
import uuid
import re

from .models import Visitor, Department, VisitorLog
from .serializers import (
    VisitorSerializer,
    VisitorCreateSerializer,
    DepartmentSerializer,
    VisitorLogSerializer,
    QRVerificationSerializer,
)
from utils.qr_generator import generate_qr_code
from utils.email_service_memory import send_visitor_confirmation_memory_only
from utils.excel_export import (
    export_visitors_to_excel_memory,
    create_bulk_visitor_template,
    get_predefined_hosts,
)


def is_weekday(visit_date):
    """Check if the given date is a weekday (Monday-Friday)"""
    return visit_date.weekday() < 5


def is_business_hours(visit_time):
    """Check if the given time is within business hours (9 AM - 5 PM)"""
    business_start = time(9, 0)
    business_end = time(17, 0)
    return business_start <= visit_time <= business_end


def get_hour_slot(visit_time):
    """Get the hour slot for the given time (e.g., 9:30 -> 9, 14:45 -> 14)"""
    return visit_time.hour


def check_slot_availability(visit_date, visit_time, exclude_visitor_id=None):
    """Check if a slot is available for the given date and time"""
    from django.utils import timezone

    now = timezone.now()
    current_date = now.date()
    current_time = now.time()

    if visit_date == current_date:
        # Add 30 minutes buffer to current time
        from datetime import timedelta

        buffer_time = (
            datetime.combine(current_date, current_time) + timedelta(minutes=30)
        ).time()

        if visit_time <= buffer_time:
            return (
                False,
                "Cannot book visits for past times or too close to current time. Please select a time at least 30 minutes from now.",
            )

    # If visit date is in the past
    if visit_date < current_date:
        return False, "Cannot book visits for past dates. Please select a future date."

    # Check if it's a weekday
    if not is_weekday(visit_date):
        return False, "Visits are only allowed on weekdays (Monday-Friday)"

    # Check if it's within business hours
    if not is_business_hours(visit_time):
        return False, "Visits are only allowed between 9:00 AM and 5:00 PM"

    # Check if the hour slot has less than 2 bookings
    hour_slot = get_hour_slot(visit_time)
    start_time = time(hour_slot, 0)
    end_time = time(hour_slot, 59, 59)

    # Count existing bookings for this hour slot
    bookings_query = Visitor.objects.filter(
        visit_date=visit_date,
        visit_time__gte=start_time,
        visit_time__lte=end_time,
        status__in=["pending", "verified"],  # Only count active bookings
    )

    # Exclude specific visitor if provided (for rescheduling)
    if exclude_visitor_id:
        bookings_query = bookings_query.exclude(id=exclude_visitor_id)

    existing_bookings = bookings_query.count()

    if existing_bookings >= 20:
        return (
            False,
            f"Hour slot {hour_slot}:00-{hour_slot+1}:00 is fully booked (20/20 slots taken)",
        )

    return True, "Slot available"


def get_available_slots(visit_date, exclude_visitor_id=None):
    """Get all available time slots for a given date"""
    from django.utils import timezone

    if not is_weekday(visit_date):
        return []

    # Get current date and time
    now = timezone.now()
    current_date = now.date()
    current_time = now.time()

    available_slots = []
    business_hours = list(range(9, 17))  # 9 AM to 5 PM (exclusive)

    for hour in business_hours:
        start_time = time(hour, 0)
        end_time = time(hour, 59, 59)

        # Count existing bookings for this hour slot
        bookings_query = Visitor.objects.filter(
            visit_date=visit_date,
            visit_time__gte=start_time,
            visit_time__lte=end_time,
            status__in=["pending", "verified"],
        )

        # Exclude specific visitor if provided (for rescheduling)
        if exclude_visitor_id:
            bookings_query = bookings_query.exclude(id=exclude_visitor_id)

        existing_bookings = bookings_query.count()

        available_count = 20 - existing_bookings
        if available_count > 0:
            # Generate available times for this hour
            available_times = []

            # Add 30 minutes buffer to current time for today
            from datetime import timedelta

            buffer_time = (
                datetime.combine(current_date, current_time) + timedelta(minutes=30)
            ).time()

            # Check if :00 time slot is available
            if existing_bookings == 0 or (
                existing_bookings == 1
                and not bookings_query.filter(visit_time=time(hour, 0)).exists()
            ):
                time_00 = time(hour, 0)
                # Only include if it's not in the past (with buffer)
                if not (visit_date == current_date and time_00 <= buffer_time):
                    available_times.append(f"{hour:02d}:00")

            # Check if :30 time slot is available
            if existing_bookings == 0 or (
                existing_bookings == 1
                and not bookings_query.filter(visit_time=time(hour, 30)).exists()
            ):
                time_30 = time(hour, 30)
                # Only include if it's not in the past (with buffer)
                if not (visit_date == current_date and time_30 <= buffer_time):
                    available_times.append(f"{hour:02d}:30")

            # Only add slot if there are available times
            if available_times:
                available_slots.append(
                    {
                        "hour": hour,
                        "time_slot": f"{hour:02d}:00-{hour+1:02d}:00",
                        "available_count": len(available_times),
                        "times": available_times,
                    }
                )

    return available_slots


class DepartmentListView(generics.ListAPIView):
    queryset = Department.objects.all()
    serializer_class = DepartmentSerializer
    permission_classes = [AllowAny]


class VisitorListCreateView(generics.ListCreateAPIView):
    queryset = Visitor.objects.all()
    filter_backends = [DjangoFilterBackend]
    filterset_fields = ["status", "department", "purpose"]
    search_fields = ["name", "email", "company"]
    permission_classes = [AllowAny]  # Change to IsAuthenticated in production

    def get_serializer_class(self):
        if self.request.method == "POST":
            return VisitorCreateSerializer
        return VisitorSerializer

    def list(self, request, *args, **kwargs):
        # Update expired visitors before listing
        update_expired_visitors()
        return super().list(request, *args, **kwargs)

    def perform_create(self, serializer):
        # Get the visit date and time from the serializer data
        visit_date = serializer.validated_data.get("visit_date")
        visit_time = serializer.validated_data.get("visit_time")

        # Check slot availability before creating the visitor
        is_available, message = check_slot_availability(visit_date, visit_time)
        if not is_available:
            from rest_framework.exceptions import ValidationError

            raise ValidationError({"slot_availability": message})

        # Generate unique QR code
        qr_code = f"VMS-{uuid.uuid4().hex[:8].upper()}"

        # Save visitor
        visitor = serializer.save(qr_code=qr_code)

        # Send confirmation email with QR code (memory-only, no local files)
        send_visitor_confirmation_memory_only(visitor)

        # Send host notification email
        from utils.email_service_memory import send_host_notification

        send_host_notification(visitor)

        # Log the registration
        VisitorLog.objects.create(
            visitor=visitor,
            action="registered",
            user=self.request.user if self.request.user.is_authenticated else None,
        )


def update_expired_visitors():
    """Update status of visitors whose visit dates have passed"""
    from django.utils import timezone

    today = timezone.now().date()

    # Find pending visitors with past dates - mark as expired
    pending_to_expire = Visitor.objects.filter(status="pending", visit_date__lt=today)

    # Find verified visitors with past dates - mark as completed
    verified_to_complete = Visitor.objects.filter(
        status="verified", visit_date__lt=today
    )

    # Update their statuses
    expired_count = pending_to_expire.update(status="expired")
    completed_count = verified_to_complete.update(status="completed")

    if expired_count > 0:
        print(f"Updated {expired_count} pending visitors to expired status")

    if completed_count > 0:
        print(f"Updated {completed_count} verified visitors to completed status")

    return {"expired": expired_count, "completed": completed_count}


class VisitorDetailView(generics.RetrieveUpdateDestroyAPIView):
    queryset = Visitor.objects.all()
    serializer_class = VisitorSerializer
    permission_classes = [AllowAny]


@api_view(["POST"])
@permission_classes([AllowAny])
def verify_qr_code(request):
    # Debug: Print the raw request data
    print(f"Raw request data: {request.data}")
    print(f"QR code from request: '{request.data.get('qr_code', '')}'")
    print(f"QR code length: {len(request.data.get('qr_code', ''))}")
    print(f"QR code repr: {repr(request.data.get('qr_code', ''))}")

    serializer = QRVerificationSerializer(data=request.data)
    if serializer.is_valid():
        qr_code = serializer.validated_data["qr_code"]
        print(f"Validated QR code: '{qr_code}'")
        print(f"Validated QR code length: {len(qr_code)}")
        print(f"Validated QR code bytes: {qr_code.encode('utf-8')}")
    else:
        print(f"Serializer validation errors: {serializer.errors}")
        return Response(
            {
                "success": False,
                "message": "Invalid QR code format",
                "errors": serializer.errors,
            },
            status=400,
        )

    try:
        visitor = Visitor.objects.get(qr_code=qr_code, is_active=True)

        # Check if QR code has expired
        if visitor.is_qr_expired():
            return Response(
                {
                    "success": False,
                    "message": f"QR code has expired! This QR code was only valid for {visitor.visit_date}. Please contact the host to reschedule your visit.",
                    "is_expired": True,
                    "visit_date": str(visitor.visit_date),
                    "expired_at": "End of visit day",
                },
                status=status.HTTP_400_BAD_REQUEST,
            )

        # Check if today is not the visit day
        if not visitor.is_visit_day():
            return Response(
                {
                    "success": False,
                    "message": f"QR code can only be used on the visit day ({visitor.visit_date}). Please come back on the scheduled date.",
                    "is_wrong_day": True,
                    "visit_date": str(visitor.visit_date),
                    "current_date": str(timezone.now().date()),
                },
                status=status.HTTP_400_BAD_REQUEST,
            )

        # Check if visitor is already verified
        if visitor.status == "verified":
            return Response(
                {
                    "success": True,
                    "message": "Visitor already verified!",
                    "visitor": VisitorSerializer(visitor).data,
                    "already_verified": True,
                }
            )

        # Check if visitor has been rescheduled
        if visitor.is_rescheduled:
            return Response(
                {
                    "success": False,
                    "message": f"Your visit has been rescheduled! New date: {visitor.visit_date}, New time: {visitor.visit_time}. Please use the updated QR code for verification.",
                    "is_rescheduled": True,
                    "new_date": str(visitor.visit_date),
                    "new_time": str(visitor.visit_time),
                    "original_date": (
                        str(visitor.original_visit_date)
                        if visitor.original_visit_date
                        else None
                    ),
                    "original_time": (
                        str(visitor.original_visit_time)
                        if visitor.original_visit_time
                        else None
                    ),
                },
                status=status.HTTP_400_BAD_REQUEST,
            )

        # Update visitor status to verified
        visitor.status = "verified"
        visitor.checked_in_at = timezone.now()
        visitor.save()

        # Log the verification
        VisitorLog.objects.create(
            visitor=visitor,
            action="verified",
            user=request.user if request.user.is_authenticated else None,
        )

        # Schedule automatic completion at end of day
        schedule_visit_completion(visitor)

        return Response(
            {
                "success": True,
                "message": "Visitor verified successfully!",
                "visitor": VisitorSerializer(visitor).data,
                "already_verified": False,
            }
        )

    except Visitor.DoesNotExist:
        return Response(
            {
                "success": False,
                "message": "Invalid QR code. Visitor not found in records.",
            },
            status=status.HTTP_404_NOT_FOUND,
        )


@api_view(["GET"])
@permission_classes([AllowAny])
def dashboard_stats(request):
    from django.db.models import Count
    from datetime import date, timedelta

    # Update expired visitors first
    update_expired_visitors()

    today = date.today()

    # Basic stats
    stats = {
        "total_visitors_today": Visitor.objects.filter(visit_date=today).count(),
        "verified_visitors": Visitor.objects.filter(status="verified").count(),
        "pending_verification": Visitor.objects.filter(status="pending").count(),
        "total_this_month": Visitor.objects.filter(
            visit_date__month=today.month, visit_date__year=today.year
        ).count(),
        "recent_visitors": VisitorSerializer(Visitor.objects.all()[:5], many=True).data,
        "all_visitors": VisitorSerializer(
            Visitor.objects.all().order_by("-created_at"), many=True
        ).data,
    }

    # Department stats
    department_stats = (
        Visitor.objects.values("department__name")
        .annotate(count=Count("id"))
        .order_by("-count")
    )

    # Purpose stats
    purpose_stats = (
        Visitor.objects.values("purpose").annotate(count=Count("id")).order_by("-count")
    )

    return Response(
        {
            **stats,
            "department_stats": list(department_stats),
            "purpose_stats": list(purpose_stats),
        }
    )


class VisitorSearchView(generics.ListAPIView):
    serializer_class = VisitorSerializer
    permission_classes = [AllowAny]

    def get_queryset(self):
        query = self.request.query_params.get("q", "")
        if query:
            return Visitor.objects.filter(
                Q(name__icontains=query)
                | Q(email__icontains=query)
                | Q(company__icontains=query)
                | Q(qr_code__icontains=query)
            )
        return Visitor.objects.none()


@api_view(["GET"])
@permission_classes([AllowAny])
def export_visitors_excel(request):
    """Export visitors to Excel format with optional date filtering"""
    try:
        # Get date range parameters
        date_from = request.query_params.get("date_from")
        date_to = request.query_params.get("date_to")

        # Start with all visitors
        visitors = Visitor.objects.all()

        # Apply date filters if provided
        if date_from:
            try:
                from_date = datetime.strptime(date_from, "%Y-%m-%d").date()
                visitors = visitors.filter(visit_date__gte=from_date)
            except ValueError:
                return Response(
                    {"error": "Invalid date_from format. Use YYYY-MM-DD"},
                    status=status.HTTP_400_BAD_REQUEST,
                )

        if date_to:
            try:
                to_date = datetime.strptime(date_to, "%Y-%m-%d").date()
                visitors = visitors.filter(visit_date__lte=to_date)
            except ValueError:
                return Response(
                    {"error": "Invalid date_to format. Use YYYY-MM-DD"},
                    status=status.HTTP_400_BAD_REQUEST,
                )

        # Order by created_at
        visitors = visitors.order_by("-created_at")

        # Generate filename with date range if provided
        filename = "visitors_export"
        if date_from and date_to:
            filename += f"_{date_from}_to_{date_to}"
        elif date_from:
            filename += f"_from_{date_from}"
        elif date_to:
            filename += f"_to_{date_to}"
        filename += ".xlsx"

        # Generate Excel file in memory
        excel_data = export_visitors_to_excel_memory(visitors, filename)

        # Create HTTP response
        response = HttpResponse(
            excel_data,
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        response["Content-Disposition"] = f"attachment; filename={filename}"
        response["Content-Length"] = len(excel_data)

        return response

    except Exception as e:
        return Response(
            {"error": f"Failed to export visitors: {str(e)}"},
            status=status.HTTP_500_INTERNAL_SERVER_ERROR,
        )


@api_view(["POST"])
@permission_classes([AllowAny])
def resend_visitor_email(request, visitor_id):
    """Resend confirmation email to visitor"""
    try:
        visitor = get_object_or_404(Visitor, id=visitor_id)

        # Send confirmation email
        result = send_visitor_confirmation_memory_only(visitor)

        # Log the email resend
        VisitorLog.objects.create(
            visitor=visitor,
            action="email_resent",
            user=request.user if request.user.is_authenticated else None,
        )

        return Response(
            {
                "message": f"Email resent successfully to {visitor.email}",
                "visitor_name": visitor.name,
            }
        )

    except Exception as e:
        return Response(
            {"error": f"Failed to resend email: {str(e)}"},
            status=status.HTTP_500_INTERNAL_SERVER_ERROR,
        )


@api_view(["DELETE"])
@permission_classes([AllowAny])
def delete_visitor(request, visitor_id):
    """Delete a visitor"""
    try:
        visitor = get_object_or_404(Visitor, id=visitor_id)

        # Log the deletion before deleting
        VisitorLog.objects.create(
            visitor=visitor,
            action="deleted",
            user=request.user if request.user.is_authenticated else None,
        )

        # Delete the visitor
        visitor.delete()

        return Response({"message": f"Visitor {visitor.name} deleted successfully"})

    except Exception as e:
        return Response(
            {"error": f"Failed to delete visitor: {str(e)}"},
            status=status.HTTP_500_INTERNAL_SERVER_ERROR,
        )


@api_view(["GET"])
@permission_classes([AllowAny])
def check_slot_availability_api(request):
    """Check if a specific slot is available"""
    visit_date_str = request.query_params.get("visit_date")
    visit_time_str = request.query_params.get("visit_time")

    if not visit_date_str or not visit_time_str:
        return Response(
            {"error": "visit_date and visit_time parameters are required"},
            status=status.HTTP_400_BAD_REQUEST,
        )

    try:
        visit_date = datetime.strptime(visit_date_str, "%Y-%m-%d").date()
        visit_time = datetime.strptime(visit_time_str, "%H:%M").time()

        is_available, message = check_slot_availability(visit_date, visit_time)

        return Response(
            {
                "available": is_available,
                "message": message,
                "visit_date": visit_date_str,
                "visit_time": visit_time_str,
            }
        )

    except ValueError as e:
        return Response(
            {"error": f"Invalid date or time format: {str(e)}"},
            status=status.HTTP_400_BAD_REQUEST,
        )


@api_view(["GET"])
@permission_classes([AllowAny])
def get_available_slots_api(request):
    """Get all available slots for a given date"""
    visit_date_str = request.query_params.get("visit_date")
    exclude_visitor_id = request.query_params.get("exclude_visitor_id")

    if not visit_date_str:
        return Response(
            {"error": "visit_date parameter is required"},
            status=status.HTTP_400_BAD_REQUEST,
        )

    try:
        visit_date = datetime.strptime(visit_date_str, "%Y-%m-%d").date()
        available_slots = get_available_slots(visit_date, exclude_visitor_id)

        return Response(
            {
                "visit_date": visit_date_str,
                "available_slots": available_slots,
                "is_weekday": is_weekday(visit_date),
            }
        )

    except ValueError as e:
        return Response(
            {"error": f"Invalid date format: {str(e)}"},
            status=status.HTTP_400_BAD_REQUEST,
        )


@api_view(["POST"])
@permission_classes([AllowAny])
def reschedule_visitor(request, visitor_id):
    """Reschedule a visitor's visit date and time"""
    try:
        visitor = get_object_or_404(Visitor, id=visitor_id)

        # Get new date and time from request
        new_date_str = request.data.get("visit_date")
        new_time_str = request.data.get("visit_time")

        if not new_date_str or not new_time_str:
            return Response(
                {"error": "visit_date and visit_time are required"},
                status=status.HTTP_400_BAD_REQUEST,
            )

        # Parse new date and time
        new_date = datetime.strptime(new_date_str, "%Y-%m-%d").date()
        new_time = datetime.strptime(new_time_str, "%H:%M").time()

        # Check if trying to reschedule to the same date and time
        if visitor.visit_date == new_date and visitor.visit_time == new_time:
            return Response(
                {"error": "Cannot reschedule to the same date and time"},
                status=status.HTTP_400_BAD_REQUEST,
            )

        # Check slot availability for new time (excluding current visitor's slot)
        is_available, message = check_slot_availability(
            new_date, new_time, exclude_visitor_id=visitor.id
        )
        if not is_available:
            return Response({"error": message}, status=status.HTTP_400_BAD_REQUEST)

        # Store old values for logging and reschedule tracking
        old_date = visitor.visit_date
        old_time = visitor.visit_time

        # Generate new QR code for rescheduled visit
        import uuid

        new_qr_code = str(uuid.uuid4())[:8].upper()

        # Update visitor with new date, time, and QR code
        visitor.visit_date = new_date
        visitor.visit_time = new_time
        visitor.qr_code = new_qr_code  # Generate new QR code
        visitor.is_rescheduled = True
        visitor.original_visit_date = old_date
        visitor.original_visit_time = old_time
        visitor.updated_at = timezone.now()
        visitor.save()

        # Log the reschedule
        VisitorLog.objects.create(
            visitor=visitor,
            action="rescheduled",
            user=request.user if request.user.is_authenticated else None,
            notes=f"Rescheduled from {old_date} {old_time} to {new_date} {new_time}. New QR code: {new_qr_code}",
        )

        # Send reschedule notification email with new QR code
        from utils.email_service_memory import send_reschedule_notification

        send_reschedule_notification(visitor, old_date, old_time, new_date, new_time)

        return Response(
            {
                "message": f"Visit rescheduled successfully from {old_date} {old_time} to {new_date} {new_time}",
                "visitor": VisitorSerializer(visitor).data,
                "old_date": old_date,
                "old_time": old_time,
                "new_date": new_date,
                "new_time": new_time,
            }
        )

    except ValueError as e:
        return Response(
            {"error": f"Invalid date or time format: {str(e)}"},
            status=status.HTTP_400_BAD_REQUEST,
        )
    except Exception as e:
        return Response(
            {"error": f"Failed to reschedule visitor: {str(e)}"},
            status=status.HTTP_500_INTERNAL_SERVER_ERROR,
        )


@api_view(["GET"])
@permission_classes([AllowAny])
def get_hosts(request):
    """Get list of predefined hosts for bulk registration"""
    try:
        hosts = get_predefined_hosts()
        return Response({"hosts": hosts, "total": len(hosts)})
    except Exception as e:
        return Response(
            {"error": f"Failed to get hosts: {str(e)}"},
            status=status.HTTP_500_INTERNAL_SERVER_ERROR,
        )


@api_view(["GET"])
@permission_classes([AllowAny])
def download_bulk_template(request):
    """Download Excel template for bulk visitor registration"""
    try:
        # Generate template
        template_data = create_bulk_visitor_template()

        # Create HTTP response
        response = HttpResponse(
            template_data,
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        response["Content-Disposition"] = (
            "attachment; filename=bulk_visitor_template.xlsx"
        )
        response["Content-Length"] = len(template_data)

        return response

    except Exception as e:
        return Response(
            {"error": f"Failed to generate template: {str(e)}"},
            status=status.HTTP_500_INTERNAL_SERVER_ERROR,
        )


@api_view(["POST"])
@permission_classes([AllowAny])
def bulk_upload_visitors(request):
    """Bulk upload visitors from Excel file with form data for common fields"""
    try:
        if "file" not in request.FILES:
            return Response(
                {"error": "No file provided"},
                status=status.HTTP_400_BAD_REQUEST,
            )

        file = request.FILES["file"]

        if not file.name.endswith((".xlsx", ".xls")):
            return Response(
                {"error": "File must be an Excel file (.xlsx or .xls)"},
                status=status.HTTP_400_BAD_REQUEST,
            )

        # Check if file is empty
        if file.size == 0:
            return Response(
                {
                    "error": "Uploaded file is empty. Please ensure the file contains data."
                },
                status=status.HTTP_400_BAD_REQUEST,
            )

        # Get common fields from form data
        purpose = request.data.get("purpose")
        department_id = request.data.get("department_id")
        visit_date_str = request.data.get("visit_date")
        visit_time_str = request.data.get("visit_time")
        host_name = request.data.get("host_name")
        host_email = request.data.get("host_email")

        # Validate common fields (department is now optional)
        if not all([purpose, visit_date_str, visit_time_str, host_name, host_email]):
            return Response(
                {
                    "error": "Missing required form fields: purpose, visit_date, visit_time, host_name, host_email"
                },
                status=status.HTTP_400_BAD_REQUEST,
            )

        # Validate purpose
        valid_purposes = [
            "business_meeting",
            "interview",
            "delivery",
            "maintenance",
            "training",
            "i_factory_visit",
            "i_factory_training",
            "other",
        ]
        if purpose not in valid_purposes:
            return Response(
                {
                    "error": f"Invalid purpose. Must be one of: {', '.join(valid_purposes)}"
                },
                status=status.HTTP_400_BAD_REQUEST,
            )

        # Validate department (optional now)
        department = None
        if department_id:
            try:
                dept_id = int(department_id)
                department = Department.objects.get(id=dept_id)
            except (ValueError, Department.DoesNotExist):
                return Response(
                    {"error": "Invalid department ID"},
                    status=status.HTTP_400_BAD_REQUEST,
                )

        # Validate and parse visit date and time
        try:
            visit_date = datetime.strptime(visit_date_str, "%Y-%m-%d").date()
            visit_time = datetime.strptime(visit_time_str, "%H:%M").time()
        except ValueError:
            return Response(
                {
                    "error": "Invalid date or time format. Use YYYY-MM-DD for date and HH:MM for time"
                },
                status=status.HTTP_400_BAD_REQUEST,
            )

        # Validate visit date and time
        if visit_date < date.today():
            return Response(
                {"error": "Visit date cannot be in the past"},
                status=status.HTTP_400_BAD_REQUEST,
            )

        if not is_weekday(visit_date):
            return Response(
                {"error": "Visit date must be a weekday (Monday-Friday)"},
                status=status.HTTP_400_BAD_REQUEST,
            )

        if not is_business_hours(visit_time):
            return Response(
                {"error": "Visit time must be between 9:00 AM and 5:00 PM"},
                status=status.HTTP_400_BAD_REQUEST,
            )

        # Validate host email
        if not re.match(
            r"^[a-zA-Z0-9._%+-]+@[a-zA-Z0-9.-]+\.[a-zA-Z]{2,}$", host_email
        ):
            return Response(
                {"error": "Invalid host email format"},
                status=status.HTTP_400_BAD_REQUEST,
            )

        import openpyxl
        from io import BytesIO

        # Read Excel file
        try:
            file_content = file.read()
            if len(file_content) == 0:
                return Response(
                    {
                        "error": "Uploaded file is empty or corrupted. Please ensure the file contains valid Excel data."
                    },
                    status=status.HTTP_400_BAD_REQUEST,
                )
            workbook = openpyxl.load_workbook(BytesIO(file_content))
            worksheet = workbook.active

            # Check if worksheet has data
            if worksheet.max_row < 2:
                return Response(
                    {
                        "error": "Excel file must contain at least a header row and one data row."
                    },
                    status=status.HTTP_400_BAD_REQUEST,
                )
        except Exception as e:
            import traceback

            traceback.print_exc()
            return Response(
                {
                    "error": f"Failed to read Excel file. Please ensure it's a valid Excel file (.xlsx or .xls). Error: {str(e)}"
                },
                status=status.HTTP_400_BAD_REQUEST,
            )

        # Get headers from first row and normalize them
        headers = []
        for cell in worksheet[1]:
            if cell.value:
                headers.append(str(cell.value).strip())
            else:
                headers.append("")

        # Expected headers for visitor-specific data only
        expected_headers = [
            "Name*",
            "Email*",
            "Phone*",
            "Company/Organization*",
            "Visitor Type*",
            "Visitor Category*",
        ]

        # Validate headers presence - improved matching logic
        header_map = {}  # Maps expected header to actual header in Excel

        # Create a list of available headers for debugging
        available_headers = [h for h in headers if h]

        for expected in expected_headers:
            found = False
            # Try exact match first
            if expected in headers:
                header_map[expected] = expected
                found = True
            else:
                # Try case-insensitive match without extra spaces
                for actual in headers:
                    # Remove asterisks and extra spaces for comparison
                    expected_normalized = expected.lower().strip()
                    actual_normalized = actual.lower().strip()

                    if expected_normalized == actual_normalized:
                        header_map[expected] = actual
                        found = True
                        break

            if not found:
                import traceback

                return Response(
                    {
                        "error": f"Invalid template format. Missing required column: '{expected}'. Available columns: {available_headers}"
                    },
                    status=status.HTTP_400_BAD_REQUEST,
                )

        visitors_data = []
        errors = []
        warnings = []

        # Process each row
        for row_num in range(2, worksheet.max_row + 1):
            row_data = {}
            row_errors = []

            # Read cell values
            for col_num, header in enumerate(headers, 1):
                if not header:
                    continue
                cell_value = worksheet.cell(row=row_num, column=col_num).value
                if cell_value is not None:
                    row_data[header] = str(cell_value).strip()
                else:
                    row_data[header] = ""

            # Skip empty rows
            if not any(row_data.values()):
                continue

            # Validate required fields using mapping
            for field in expected_headers:
                actual_header = header_map[field]
                if not row_data.get(actual_header):
                    row_errors.append(f"{field} is required")

            # Validate email format
            email_field = header_map["Email*"]
            if row_data.get(email_field) and not re.match(
                r"^[a-zA-Z0-9._%+-]+@[a-zA-Z0-9.-]+\.[a-zA-Z]{2,}$",
                row_data[email_field],
            ):
                row_errors.append("Invalid email format")

            # Validate and map visitor type
            type_field = header_map["Visitor Type*"]
            raw_type = row_data.get(type_field, "").lower().strip()
            if raw_type in ["emp", "employee", "professional", "prof"]:
                row_data["MappedType"] = "professional"
            elif raw_type == "student":
                row_data["MappedType"] = "student"
            elif raw_type:
                row_data["MappedType"] = (
                    raw_type  # Keep as is, will fail model-level if invalid but let's be flexible
                )
                warnings.append(
                    f"Row {row_num}: Non-standard visitor type '{raw_type}' used."
                )

            # Validate and map visitor category
            cat_field = header_map["Visitor Category*"]
            raw_cat = row_data.get(cat_field, "").lower().strip()
            if raw_cat in ["industry", "industrial"]:
                row_data["MappedCategory"] = "industry"
            elif raw_cat in ["academic", "academy", "education"]:
                row_data["MappedCategory"] = "academic"
            elif raw_cat in ["government", "govt"]:
                row_data["MappedCategory"] = "government"
            elif raw_cat in ["other"]:
                row_data["MappedCategory"] = "other"
            elif raw_cat:
                row_data["MappedCategory"] = raw_cat
                warnings.append(
                    f"Row {row_num}: Non-standard category '{raw_cat}' used."
                )

            if row_errors:
                errors.append({"row": row_num, "data": row_data, "errors": row_errors})
            else:
                visitors_data.append({"row": row_num, "data": row_data})

        # If there are validation errors, return them
        if errors:
            return Response(
                {
                    "success": False,
                    "message": f"Found {len(errors)} rows with errors",
                    "errors": errors[:10],
                    "total_errors": len(errors),
                    "processed_rows": 0,
                },
                status=status.HTTP_400_BAD_REQUEST,
            )

        # Check if there are any visitors to process
        if not visitors_data:
            return Response(
                {
                    "error": "No valid visitor data found in Excel file. Please ensure the file has at least one row of visitor data."
                },
                status=status.HTTP_400_BAD_REQUEST,
            )

        # Check slot availability for all visitors (they all have the same visit date/time)
        total_slots_needed = len(visitors_data)

        # Validate against maximum upload limit
        MAX_BULK_UPLOAD = 20
        if total_slots_needed > MAX_BULK_UPLOAD:
            return Response(
                {
                    "error": f"Cannot upload more than {MAX_BULK_UPLOAD} visitors at once. You are trying to upload {total_slots_needed} visitors."
                },
                status=status.HTTP_400_BAD_REQUEST,
            )

        # Check actual visitor capacity for the hour slot
        hour_slot = get_hour_slot(visit_time)
        start_time = time(hour_slot, 0)
        end_time = time(hour_slot, 59, 59)

        # Count existing bookings for this hour slot
        existing_bookings = Visitor.objects.filter(
            visit_date=visit_date,
            visit_time__gte=start_time,
            visit_time__lte=end_time,
            status__in=["pending", "verified"],
        ).count()

        # Maximum visitors per hour slot
        MAX_VISITORS_PER_HOUR = 20
        available_capacity = MAX_VISITORS_PER_HOUR - existing_bookings

        if available_capacity < total_slots_needed:
            return Response(
                {
                    "error": f"Not enough capacity available. Need {total_slots_needed} slots but only {available_capacity} available for {visit_date} at {visit_time}. Hour slot {hour_slot}:00-{hour_slot+1}:00 has {existing_bookings}/{MAX_VISITORS_PER_HOUR} slots booked."
                },
                status=status.HTTP_400_BAD_REQUEST,
            )

        created_visitors = []
        failed_visitors = []
        visitor_ids_for_email = []

        # Bulk create visitors for better performance
        visitors_to_create = []
        visitor_logs_to_create = []

        for visitor_info in visitors_data:
            try:
                row_data = visitor_info["data"]
                row_num = visitor_info["row"]

                # Prepare visitor object (don't save yet)
                visitor = Visitor(
                    name=row_data.get(header_map["Name*"]),
                    email=row_data.get(header_map["Email*"]),
                    phone=row_data.get(header_map["Phone*"]),
                    company=row_data.get(header_map["Company/Organization*"], ""),
                    visitor_type=row_data.get("MappedType", "professional"),
                    visitor_category=row_data.get("MappedCategory", "industry"),
                    purpose=purpose,
                    department=department,
                    visit_date=visit_date,
                    visit_time=visit_time,
                    host_name=host_name,
                    host_email=host_email,
                    notes=row_data.get("Notes", ""),
                    qr_code=f"VMS-{uuid.uuid4().hex[:8].upper()}",
                    status="pending",
                )
                visitors_to_create.append(visitor)

            except Exception as e:
                failed_visitors.append(
                    {"row": row_num, "data": row_data, "error": str(e)}
                )

        # Bulk create all visitors at once (much faster than individual creates)
        if visitors_to_create:
            created_visitor_objects = Visitor.objects.bulk_create(visitors_to_create)

            # Prepare visitor logs for bulk creation
            for idx, visitor in enumerate(created_visitor_objects):
                row_num = visitors_data[idx]["row"]

                visitor_logs_to_create.append(
                    VisitorLog(
                        visitor=visitor,
                        action="registered",
                        user=request.user if request.user.is_authenticated else None,
                        notes=f"Bulk registered from Excel row {row_num}",
                    )
                )

                created_visitors.append(
                    {
                        "id": str(visitor.id),
                        "name": visitor.name,
                        "email": visitor.email,
                        "qr_code": visitor.qr_code,
                        "visit_date": str(visitor.visit_date),
                        "visit_time": str(visitor.visit_time),
                        "row": row_num,
                    }
                )

                # Collect visitor IDs for async email sending
                visitor_ids_for_email.append(str(visitor.id))

            # Bulk create visitor logs
            VisitorLog.objects.bulk_create(visitor_logs_to_create)

            # Send emails asynchronously using Celery (non-blocking)
            try:
                from visitors.tasks import (
                    send_visitor_email_async,
                    send_bulk_host_notification_async,
                )

                # Queue individual visitor emails asynchronously
                for visitor_id in visitor_ids_for_email:
                    send_visitor_email_async.delay(visitor_id)

                # Queue bulk host notification asynchronously
                send_bulk_host_notification_async.delay(
                    host_name=host_name,
                    host_email=host_email,
                    visitors=created_visitors,
                    visit_date=str(visit_date),
                    visit_time=str(visit_time),
                    purpose=purpose,
                )

                warnings.append(
                    "Emails are being sent in the background. Visitors will receive their confirmation emails shortly."
                )

            except Exception as e:
                # If Celery is not available, fall back to synchronous email sending
                warnings.append(
                    f"Async email service unavailable. Emails will be sent synchronously: {str(e)}"
                )

                # Fallback: send emails synchronously
                for visitor_id in visitor_ids_for_email:
                    try:
                        visitor = Visitor.objects.get(id=visitor_id)
                        send_visitor_confirmation_memory_only(visitor)
                    except Exception as email_error:
                        warnings.append(
                            f"Failed to send email to visitor {visitor_id}: {str(email_error)}"
                        )

                # Send bulk host notification synchronously
                try:
                    from utils.email_service_memory import send_bulk_host_notification

                    send_bulk_host_notification(
                        host_name=host_name,
                        host_email=host_email,
                        visitors=created_visitors,
                        visit_date=visit_date,
                        visit_time=visit_time,
                        purpose=purpose,
                    )
                except Exception as e:
                    warnings.append(f"Failed to send summary email to host: {str(e)}")

        return Response(
            {
                "success": True,
                "message": f"Successfully processed {len(created_visitors)} visitors",
                "created_visitors": created_visitors,
                "failed_visitors": failed_visitors,
                "warnings": warnings,
                "total_processed": len(visitors_data),
                "successful": len(created_visitors),
                "failed": len(failed_visitors),
                "visit_details": {
                    "visit_date": str(visit_date),
                    "visit_time": str(visit_time),
                    "purpose": purpose,
                    "host_name": host_name,
                    "host_email": host_email,
                },
            }
        )

    except Exception as e:
        import traceback

        traceback.print_exc()
        return Response(
            {"error": f"Failed to process bulk upload: {str(e)}"},
            status=status.HTTP_500_INTERNAL_SERVER_ERROR,
        )


def schedule_visit_completion(visitor):
    """Schedule automatic completion of visit at end of day"""
    from django.core.management import call_command
    from threading import Timer

    # Calculate time until end of day
    now = timezone.now()
    end_of_day = timezone.make_aware(
        datetime.combine(visitor.visit_date, time(23, 59, 59))
    )

    # If it's already past end of day, complete immediately
    if now >= end_of_day:
        complete_visit(visitor)
    else:
        # Schedule completion at end of day
        delay_seconds = (end_of_day - now).total_seconds()

        def complete_visit_delayed():
            complete_visit(visitor)

        # Use a timer to complete the visit at end of day
        timer = Timer(delay_seconds, complete_visit_delayed)
        timer.start()


def complete_visit(visitor):
    """Mark visit as completed and log the action"""
    if visitor.status == "verified":
        visitor.status = "completed"
        visitor.checked_out_at = timezone.now()
        visitor.save()

        # Log the completion
        VisitorLog.objects.create(
            visitor=visitor,
            action="checked_out",
            notes="Visit automatically completed at end of day",
        )


@api_view(["POST"])
@permission_classes([IsAuthenticated])
def manual_complete_visit(request, visitor_id):
    try:
        visitor = get_object_or_404(Visitor, id=visitor_id)

        if visitor.status not in ["verified", "pending"]:
            return Response(
                {
                    "error": f"Visit cannot be completed. Current status: {visitor.status}"
                },
                status=status.HTTP_400_BAD_REQUEST,
            )

        visitor.status = "completed"
        visitor.checked_out_at = timezone.now()
        visitor.save()

        # Log the manual completion
        VisitorLog.objects.create(
            visitor=visitor,
            action="checked_out",
            user=request.user,
            notes="Visit manually completed by admin",
        )

        return Response(
            {
                "success": True,
                "message": f"Visit for {visitor.name} completed successfully",
                "visitor": VisitorSerializer(visitor).data,
            }
        )

    except Exception as e:
        return Response(
            {"error": f"Failed to complete visit: {str(e)}"},
            status=status.HTTP_500_INTERNAL_SERVER_ERROR,
        )
